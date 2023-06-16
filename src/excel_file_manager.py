from file_manager import FileManager
from vacancy import Vacancy
from openpyxl import Workbook, load_workbook


class ExcelFileManager(FileManager):
    '''
    Класс для работы с Excel-файлами
    '''

    def save_to_file(self, filename):
        '''
        Функция сохраняет созданные вакансии на основе
        класса Vacancy в файл
        :param filename: название файла
        '''
        with open(f'files/{filename}.xlsx', 'w', encoding='UTF-8') as file:
            work_book = Workbook()
            sheet = work_book.active
            sheet.append([
                'название', 'описание', 'место', 'мин.зарплата', 'макс.зарплата',
                'зарплата', 'опыт', 'работадатель', 'занятость', 'адрес'
            ])

            for vacancy in Vacancy.vacancies:
                sheet.append([
                    vacancy.name, vacancy.description, vacancy.area,
                    vacancy.salary_from, vacancy.salary_to, vacancy.salary,
                    vacancy.experience, vacancy.employer, vacancy.employment,
                    vacancy.adress
                ])

            work_book.save(f'files/{filename}.xlsx')
            work_book.close()

    def load_from_file(self, filename):
        '''
        Функция загружает вакансии из файла, создавая
        объекты на основе класса Vacancy
        :param filename: название файла
        '''
        Vacancy.vacancies.clear()
        try:
            with open(f'files/{filename}.xlsx', 'r', encoding='UTF-8') as file:
                work_book = load_workbook(f'files/{filename}.xlsx')
                sheet = work_book.active

                for row in sheet.iter_rows(values_only=True):
                    row = list(row)

                    # проверяет является ли ряд названием колонок
                    if row[0] == 'название':
                        continue

                    # проверяет указана ли зарплата для установки валюты
                    if row[5] != 'Зарплата не указана':
                        row[5] = row[5].split()[1]
                    else:
                        row[5] = 'RUR'

                    Vacancy(
                        row[0],
                        row[1],
                        row[2],
                        row[3],
                        row[4],
                        row[5],
                        row[6],
                        row[7],
                        row[8],
                        row[9]
                    )
        except FileNotFoundError:
            print('Такого файла не существует!')

    def load_vacancies_by_salary(self, filename, value):
        '''
        Функция загружает вакансии из файла по переданой
        пользователем зарплата, создавая объекты на основе
        класса Vacancy
        :param value: зарплата
        :param filename: название файла
        :return: вакансии
        '''
        Vacancy.vacancies.clear()
        try:
            with open(f'files/{filename}.xlsx', 'r', encoding='UTF-8') as file:
                work_book = load_workbook(f'files/{filename}.xlsx')
                sheet = work_book.active
                user_salary = value.split('-')

                if len(user_salary) == 2 and user_salary[0].isdigit() and user_salary[1].isdigit():

                    for row in sheet.iter_rows(values_only=True):
                        row = list(row)

                        # проверяет является ли ряд названием колонок
                        if row[0] == 'название':
                            continue

                        # проверяет указана ли зарплата для установки валюты
                        if row[5] == 'Зарплата не указана':
                            continue
                        else:
                            salary = row[5].split()[0]
                            row[5] = row[5].split()[1]

                        if int(user_salary[0]) <= int(salary) <= int(user_salary[1]):
                            Vacancy(
                                row[0],
                                row[1],
                                row[2],
                                row[3],
                                row[4],
                                row[5],
                                row[6],
                                row[7],
                                row[8],
                                row[9]
                            )
                else:
                    print('Вы ввели неверные данные!')
        except FileNotFoundError:
            print('Такого файла не существует!')

    def load_vacancies_by_area(self, filename, value):
        '''
        Функция загружает вакансии из файла по переданой
        пользователем области поиска, создавая объекты на основе
        класса Vacancy
        :param value: область поиска
        :param filename: название файла
        :return: вакансии
        '''
        Vacancy.vacancies.clear()
        try:
            with open(f'files/{filename}.xlsx', 'r', encoding='UTF-8') as file:
                work_book = load_workbook(f'files/{filename}.xlsx')
                sheet = work_book.active
                user_area = value

                for row in sheet.iter_rows(values_only=True):
                    row = list(row)

                    # проверяет является ли ряд названием колонок
                    if row[0] == 'название':
                        continue

                    # проверяет указана ли зарплата для установки валюты
                    if row[5] != 'Зарплата не указана':
                        row[5] = row[5].split()[1]
                    else:
                        row[5] = 'RUR'

                    if user_area.lower() == row[2].lower():
                        Vacancy(
                            row[0],
                            row[1],
                            row[2],
                            row[3],
                            row[4],
                            row[5],
                            row[6],
                            row[7],
                            row[8],
                            row[9]
                        )

        except FileNotFoundError:
            print('Такого файла не существует!')

    def delete_vacancy(self, filename, value):
        '''
        Удаляет вакансию из файла по переданому
        номеру вакансии
        :param value: номер вакансии
        :param filename: название файла
        '''
        try:
            with open(f'files/{filename}.xlsx', 'r', encoding='UTF-8') as file:
                work_book = load_workbook(f'files/{filename}.xlsx')
                sheet = work_book.active
                try:
                    if sheet.max_row - 1 < value:
                        print('Вакансии с таким номером не существует')
                    else:
                        sheet.delete_rows(value + 1)
                except ValueError:
                    print('Вы ввели некоректные данные')
                work_book.save(f'files/{filename}.xlsx')
                work_book.close()
        except FileNotFoundError:
            print('Такого файла не существует!')
